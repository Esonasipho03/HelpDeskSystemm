from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from tickets.models import Announcement, Ticket
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Case, IntegerField, Q, Value, When
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from tickets.forms import TicketCommentForm, TicketCreateForm
from tickets.models import (
    Announcement,
    Ticket,
    TicketPriority,
    TicketStatus,
)
from django.db.models import Count, Avg, Q, F, ExpressionWrapper, DurationField
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def employee_dashboard(request):

    return render(
        request,
        "dashboard/employee_dashboard.html"
    )
    
    
@login_required
def employee_dashboard(request):
    user = request.user

    my_tickets = Ticket.objects.filter(created_by=user)

    context = {
        "open_tickets_count": my_tickets.filter(
            status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS]
        ).count(),
        "pending_reply_count": my_tickets.filter(status=TicketStatus.PENDING).count(),
        "resolved_tickets_count": my_tickets.filter(
            status=TicketStatus.RESOLVED
        ).count(),
        "recent_tickets": my_tickets[:5],
        "ticket_to_rate": my_tickets.filter(
            status=TicketStatus.RESOLVED,
            satisfaction_rating__isnull=True,
        ).first(),
        "my_assets": user.assets.all(),
        "announcements": Announcement.objects.filter(is_active=True)[:5],
    }
    return render(request, "dashboard/employee_dashboard.html", context)


@login_required
def create_ticket(request):
    initial = {}
    category = request.GET.get("category")
    if category:
        initial["category"] = category

    if request.method == "POST":
        form = TicketCreateForm(request.POST, user=request.user)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.save()
            messages.success(request, f"Ticket #{ticket.pk} submitted. We'll be in touch shortly.")
            return redirect("ticket_detail", pk=ticket.pk)
    else:
        form = TicketCreateForm(user=request.user, initial=initial)

    return render(request, "create_ticket.html", {"form": form})


@login_required
def ticket_history(request):
    tickets = Ticket.objects.filter(created_by=request.user)

    status_filter = request.GET.get("status")
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    return render(request, "ticket_history.html", {
        "tickets": tickets,
        "status_filter": status_filter,
        "status_choices": TicketStatus.choices,
    })


@login_required
def ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk, created_by=request.user)

    if request.method == "POST":
        comment_form = TicketCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.ticket = ticket
            comment.author = request.user
            comment.save()
            return redirect("ticket_detail", pk=ticket.pk)
    else:
        comment_form = TicketCommentForm()

    # requester never sees internal notes
    visible_comments = ticket.comments.filter(is_internal_note=False)

    return render(request, "ticket_detail.html", {
        "ticket": ticket,
        "comments": visible_comments,
        "comment_form": comment_form,
    })


@login_required
def rate_ticket(request, pk):
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, pk=pk, created_by=request.user)
        rating = request.POST.get("rating")
        if rating and rating.isdigit() and 1 <= int(rating) <= 5:
            ticket.satisfaction_rating = int(rating)
            ticket.save(update_fields=["satisfaction_rating"])
            messages.success(request, "Thanks for the feedback!")
    return redirect("employee_dashboard")


# ---------------------------------------------------------------------
# Technician views
# ---------------------------------------------------------------------

# Urgent first. Used to sort queues by actual urgency rather than
# alphabetically (which would put "high" before "low" before "medium"
# before "urgent" — wrong order).
PRIORITY_RANK = Case(
    When(priority=TicketPriority.CRITICAL, then=Value(0)),
    When(priority=TicketPriority.HIGH, then=Value(1)),
    When(priority=TicketPriority.MEDIUM, then=Value(2)),
    When(priority=TicketPriority.LOW, then=Value(3)),
    output_field=IntegerField(),
)

OPEN_STATUSES = [TicketStatus.OPEN, TicketStatus.PENDING, TicketStatus.IN_PROGRESS]


def _is_technician(user):
    # Adjust this to match however roles are actually modeled on your User.
    return user.is_staff or getattr(user, "role", "").lower() in {"technician", "admin"}


from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from tickets.models import Ticket, Announcement


@login_required
def technician_dashboard(request):

    user = request.user

    assigned = Ticket.objects.filter(
        assigned_to=user
    )

    context = {

        # Dashboard statistics

        "assigned_count": assigned.count(),

        "open_count": assigned.filter(
            status=TicketStatus.OPEN
        ).count(),

        "progress_count": assigned.filter(
            status=TicketStatus.IN_PROGRESS
        ).count(),

        "resolved_count": assigned.filter(
            status=TicketStatus.RESOLVED
        ).count(),

        "unassigned_count": Ticket.objects.filter(
            assigned_to__isnull=True
        ).count(),

        # Tables

        "my_queue": assigned.order_by("-updated_at")[:8],

        "urgent_tickets": Ticket.objects.filter(
            priority=TicketPriority.HIGH
        ).order_by("-created_at")[:6],

        "recent_activity": Ticket.objects.order_by(
            "-updated_at"
        )[:12],

        "announcements": Announcement.objects.filter(
            is_active=True
        )[:5],
    }

    return render(
        request,
        "dashboard/technician_dashboard.html",
        context,
    )
@login_required
def my_queue(request):
    tickets = Ticket.objects.filter(assigned_to=request.user)

    status_filter = request.GET.get("status")
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    tickets = tickets.annotate(priority_rank=PRIORITY_RANK).order_by("priority_rank", "created_at")

    return render(request, "dashboard/my_queue.html", {
        "tickets": tickets,
        "status_filter": status_filter,
        "status_choices": TicketStatus.choices,
    })


@login_required
def unassigned_queue(request):
    tickets = Ticket.objects.filter(
        assigned_to__isnull=True,
        status__in=OPEN_STATUSES,
    ).annotate(priority_rank=PRIORITY_RANK).order_by("priority_rank", "created_at")

    return render(request, "dashboard/unassigned_queue.html", {"tickets": tickets})


@login_required
def all_tickets(request):
    tickets = Ticket.objects.all().order_by("-updated_at")

    status_filter = request.GET.get("status")
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    return render(request, "dashboard/all_tickets.html", {
        "tickets": tickets,
        "status_filter": status_filter,
        "status_choices": TicketStatus.choices,
    })


@login_required
def ticket_search(request):
    query = request.GET.get("q", "").strip()
    tickets = Ticket.objects.none()

    if query:
        search_filter = (
            Q(title__icontains=query)
            | Q(description__icontains=query)
            | Q(created_by__first_name__icontains=query)
            | Q(created_by__last_name__icontains=query)
        )
        ticket_number = query.lstrip("#")
        if ticket_number.isdigit():
            search_filter |= Q(pk=int(ticket_number))

        tickets = Ticket.objects.filter(search_filter).order_by("-updated_at")

    return render(request, "dashboard/ticket_search.html", {
        "tickets": tickets,
        "query": query,
    })


@login_required
def claim_ticket(request, pk):
    """Technician assigns an unassigned ticket to themselves."""
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, pk=pk, assigned_to__isnull=True)
        ticket.assigned_to = request.user
        if ticket.status == TicketStatus.OPEN:
            ticket.status = TicketStatus.IN_PROGRESS
        ticket.save(update_fields=["assigned_to", "status"])
        messages.success(request, f"Ticket #{ticket.pk} assigned to you.")

    next_url = request.POST.get("next") or "technician_dashboard"
    return redirect(next_url)


@login_required
def update_ticket_status(request, pk):
    """Technician changes a ticket's status from the queue or detail view."""
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, pk=pk, assigned_to=request.user)
        new_status = request.POST.get("status")
        valid_statuses = {choice.value for choice in TicketStatus}
        if new_status in valid_statuses:
            ticket.status = new_status
            if new_status in {TicketStatus.RESOLVED}:
                ticket.resolved_at = timezone.now()
            ticket.save()
            messages.success(request, f"Ticket #{ticket.pk} marked as {ticket.get_status_display()}.")

    next_url = request.POST.get("next") or "my_queue"
    return redirect(next_url)


@login_required
def technician_profile(request):
    return render(request, "dashboard/technician_profile.html", {})




def _is_manager(user):
    """
    Adjust this to match your real permission model.
    Falls back to Django's built-in is_staff flag if you don't yet
    have a dedicated MANAGER role on your user/role field.
    """
    role = str(getattr(user, "role", "")).upper()
    return user.is_staff or role == "MANAGER"


def _build_technician_report_rows():
    """Shared by both the HTML view and the Excel export, so the
    two can never drift out of sync with each other."""

    technicians = User.objects.filter(
        tickets_assigned__isnull=False
    ).distinct().annotate(
        assigned_count=Count("tickets_assigned", distinct=True),
        resolved_count=Count(
            "tickets_assigned",
            filter=Q(tickets_assigned__status=TicketStatus.RESOLVED),
            distinct=True,
        ),
        closed_count=Count(
            "tickets_assigned",
            filter=Q(tickets_assigned__status=TicketStatus.RESOLVED),
            distinct=True,
        ),
        open_count=Count(
            "tickets_assigned",
            filter=Q(tickets_assigned__status__in=[
                TicketStatus.OPEN,
                TicketStatus.ASSIGNED,
                TicketStatus.IN_PROGRESS,
                TicketStatus.PENDING,
            ]),
            distinct=True,
        ),
        avg_rating=Avg("tickets_assigned__satisfaction_rating"),
    ).order_by("-resolved_count")

    # Average resolution time needs its own query — averaging a duration
    # expression alongside multiple Count(...) annotations on the same
    # join can distort both, so it's computed separately and merged below.
    resolution_durations = (
        Ticket.objects.filter(assigned_to__isnull=False, resolved_at__isnull=False)
        .annotate(
            resolution_time=ExpressionWrapper(
                F("resolved_at") - F("created_at"),
                output_field=DurationField(),
            )
        )
        .values("assigned_to")
        .annotate(avg_resolution=Avg("resolution_time"))
    )
    avg_resolution_by_tech = {
        row["assigned_to"]: row["avg_resolution"] for row in resolution_durations
    }

    rows = []
    for tech in technicians:
        avg_duration = avg_resolution_by_tech.get(tech.id)
        avg_hours = round(avg_duration.total_seconds() / 3600, 1) if avg_duration else None

        completion_rate = 0
        if tech.assigned_count:
            completion_rate = round(
                ((tech.resolved_count + tech.closed_count) / tech.assigned_count) * 100, 1
            )

        rows.append({
            "technician": tech,
            "name": tech.get_full_name() or tech.username,
            "assigned_count": tech.assigned_count,
            "resolved_count": tech.resolved_count,
            "closed_count": tech.closed_count,
            "open_count": tech.open_count,
            "avg_rating": round(tech.avg_rating, 2) if tech.avg_rating else None,
            "avg_resolution_hours": avg_hours,
            "completion_rate": completion_rate,
        })

    return rows


@login_required
def technician_efficiency_report(request):

    if not _is_manager(request.user):
        messages.error(request, "You don't have permission to view this report.")
        return redirect("technician_dashboard")

    report_rows = _build_technician_report_rows()

    return render(
        request,
        "tickets/technician_efficiency_report.html",
        {"report_rows": report_rows},
    )


@login_required
def technician_efficiency_report_export(request):

    if not _is_manager(request.user):
        messages.error(request, "You don't have permission to view this report.")
        return redirect("technician_dashboard")

    rows = _build_technician_report_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Technician Efficiency"

    headers = [
        "Technician",
        "Assigned",
        "Resolved",
        "Closed",
        "Currently Open",
        "Completion Rate (%)",
        "Avg Resolution Time (hrs)",
        "Avg Satisfaction Rating",
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="F17807", end_color="F17807", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_num, row in enumerate(rows, start=2):
        values = [
            row["name"],
            row["assigned_count"],
            row["resolved_count"],
            row["closed_count"],
            row["open_count"],
            row["completion_rate"] / 100 if row["completion_rate"] else 0,
            row["avg_resolution_hours"] if row["avg_resolution_hours"] is not None else "-",
            row["avg_rating"] if row["avg_rating"] is not None else "-",
        ]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = body_font
            cell.border = thin_border
            if col_num == 6:
                cell.number_format = "0.0%"

    # Totals row, using real formulas so the sheet stays correct if edited
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True, size=10)

    for col_letter, col_num in (("B", 2), ("C", 3), ("D", 4), ("E", 5)):
        formula_cell = ws.cell(
            row=total_row,
            column=col_num,
            value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
        )
        formula_cell.font = Font(name="Arial", bold=True, size=10)
        formula_cell.border = thin_border

    column_widths = [26, 11, 11, 11, 15, 18, 22, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="technician_efficiency_report.xlsx"'
    wb.save(response)
    return response